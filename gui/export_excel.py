import os
import json
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from server.config import PURCHASES_DIR

def export_to_excel():
    file_paths = filedialog.askopenfilenames(
        title="Välj en eller flera köp-filer",
        filetypes=[("JSON-filer", "*.json")],
        initialdir=PURCHASES_DIR
    )

    if not file_paths:
        return

    purchases = []
    for path in file_paths:
        with open(path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                if isinstance(data, list):
                    purchases.extend(data)
            except Exception as e:
                messagebox.showwarning("Fel", f"Kunde inte läsa {os.path.basename(path)}:\n{e}")

    if not purchases:
        messagebox.showinfo("Tom export", "Inga giltiga poster att exportera.")
        return

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Köp export"

    headers = ["Vara", "Stånd", "Antal", "Pris (kr)", "Totalt (kr)", "Tidpunkt"]
    ws.append(headers)

    for p in purchases:
        row = [
            p.get("name", "Okänd"),
            p.get("stand", "Okänt"),
            p.get("quantity", 0),
            p.get("price", 0.0),
            round(p.get("price", 0.0) * p.get("quantity", 0), 2),
            p.get("timestamp", "")
        ]
        ws.append(row)

    # Auto-adjust column width
    for col in ws.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    # Save to Desktop
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    filename = os.path.join(desktop, f"Kop_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filename)

    messagebox.showinfo("Export klar", f"Excel-fil sparad på skrivbordet:\n{filename}")
